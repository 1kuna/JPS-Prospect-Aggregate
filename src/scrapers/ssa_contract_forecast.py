import os
import sys
import time
import datetime
import logging
import requests
import csv
import tempfile
from io import StringIO
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import glob
import pathlib
import re
import pandas as pd

# Add the parent directory to the path so we can import from src
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from src.database.db import get_session, close_session
from src.database.models import Proposal, DataSource
from src.database.download_tracker import download_tracker

# Create logs directory if it doesn't exist
logs_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'logs')
if not os.path.exists(logs_dir):
    os.makedirs(logs_dir)

# Create downloads directory if it doesn't exist
downloads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'downloads')
if not os.path.exists(downloads_dir):
    os.makedirs(downloads_dir)

# Set up logging
log_file = os.path.join(logs_dir, f'ssa_contract_forecast_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"Logging to {log_file}")

class SSAContractForecastScraper:
    """Scraper for the SSA Contract Forecast site"""
    
    def __init__(self, debug_mode=False):
        self.base_url = "https://www.ssa.gov/osdbu/contract-forecast-intro.html"
        self.source_name = "SSA Contract Forecast"
        self.debug_mode = debug_mode
        logger.info(f"Initializing scraper with debug_mode={debug_mode}")
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
    
    def setup_browser(self):
        """Set up and configure the Playwright browser"""
        logging.info("Setting up Playwright browser")
        
        # Get the downloads directory
        download_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'downloads')
        # Ensure the download directory exists
        if not os.path.exists(download_dir):
            os.makedirs(download_dir)
        logger.info(f"Setting download directory to {download_dir}")
        
        # Convert to absolute path
        download_dir_abs = os.path.abspath(download_dir)
        logger.info(f"Absolute download path: {download_dir_abs}")
        
        try:
            # Start Playwright
            self.playwright = sync_playwright().start()
            logger.info("Started Playwright")
            
            # Set browser options
            browser_type = self.playwright.chromium
            
            # Launch the browser with appropriate options
            self.browser = browser_type.launch(
                headless=not self.debug_mode,
                downloads_path=download_dir_abs,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--disable-extensions",
                    "--disable-popup-blocking",
                    "--window-size=1920,1080"
                ]
            )
            logger.info("Launched browser")
            
            # Create a new context with download options
            self.context = self.browser.new_context(
                accept_downloads=True,
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )
            logger.info("Created browser context")
            
            # Create a new page
            self.page = self.context.new_page()
            logger.info("Created page")
            
            # Set default timeout
            self.page.set_default_timeout(60000)  # 60 seconds
            
            logger.info("Playwright browser setup complete")
            return True
        except Exception as e:
            logger.error(f"Error setting up Playwright browser: {e}")
            self.cleanup_browser()
            return False
    
    def cleanup_browser(self):
        """Clean up Playwright resources"""
        logger.info("Cleaning up Playwright resources")
        try:
            if self.page:
                logger.info("Closing page")
                self.page.close()
                self.page = None
        except Exception as e:
            logger.error(f"Error closing page: {e}")
        
        try:
            if self.context:
                logger.info("Closing context")
                self.context.close()
                self.context = None
        except Exception as e:
            logger.error(f"Error closing context: {e}")
        
        try:
            if self.browser:
                logger.info("Closing browser")
                self.browser.close()
                self.browser = None
        except Exception as e:
            logger.error(f"Error closing browser: {e}")
        
        try:
            if self.playwright:
                logger.info("Stopping playwright")
                self.playwright.stop()
                self.playwright = None
        except Exception as e:
            logger.error(f"Error stopping playwright: {e}")
    
    def parse_date(self, date_str):
        """Parse a date string into a datetime object"""
        if date_str is None:
            return None
        
        # Convert to string if it's a numeric type
        if isinstance(date_str, (int, float)):
            date_str = str(date_str)
        
        # If it's empty after conversion, return None
        if not date_str or str(date_str).strip() == "":
            return None
        
        # Try different date formats
        date_formats = [
            "%m/%d/%Y",
            "%m/%d/%y",
            "%Y-%m-%d",
            "%B %d, %Y",
            "%b %d, %Y",
            "%m-%d-%Y",
            "%m-%d-%y"
        ]
        
        # Clean up the date string
        date_str = str(date_str).strip()
        
        # Try each format
        for date_format in date_formats:
            try:
                return datetime.datetime.strptime(date_str, date_format)
            except ValueError:
                continue
        
        # If we get here, none of the formats worked
        logger.warning(f"Could not parse date: {date_str}")
        return None
    
    def parse_value(self, value_str):
        """Parse a value string into a float"""
        if value_str is None:
            return None
        
        # If it's already a numeric type, return it
        if isinstance(value_str, (int, float)):
            return float(value_str)
        
        # Convert to string if needed
        value_str = str(value_str)
        
        # If it's empty after conversion, return None
        if not value_str or value_str.strip() == "":
            return None
        
        # Remove any non-numeric characters except for decimal points
        value_str = re.sub(r'[^\d.]', '', value_str)
        
        # Try to convert to float
        try:
            return float(value_str)
        except ValueError:
            logger.warning(f"Could not parse value: {value_str}")
            return None
    
    def download_forecast_document(self):
        """Download the FY25 SSA Contract Forecast document"""
        logger.info("Attempting to download the FY25 SSA Contract Forecast document")
        
        # Check if we have a valid file using the download tracker
        if download_tracker.verify_file_exists("SSA Contract Forecast", "*01072025*.xls*"):
            # Get the downloads directory
            download_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'downloads')
            
            # Find the matching files
            matching_files = list(pathlib.Path(download_dir).glob("*01072025*.xls*"))
            
            if matching_files:
                # Sort by modification time (most recent first)
                matching_files.sort(key=os.path.getmtime, reverse=True)
                latest_file = str(matching_files[0])
                logger.info(f"Found existing 2025 file, using it instead of downloading again: {latest_file}")
                return latest_file
        
        try:
            # Navigate to the base URL
            logger.info(f"Navigating to {self.base_url}")
            self.page.goto(self.base_url, wait_until="domcontentloaded")
            
            # Wait for the page to load
            logger.info("Waiting for page to load")
            self.page.wait_for_selector("body", state="visible")
            
            # Look for the FY25 SSA Contract Forecast link
            logger.info("Looking for the FY25 SSA Contract Forecast link")
            forecast_links = self.page.query_selector_all("a:has-text('FY25 SSA Contract Forecast')")
            
            if not forecast_links:
                logger.warning("Could not find the FY25 SSA Contract Forecast link")
                # Try looking for it in the main menu
                forecast_links = self.page.query_selector_all("div.mainMenu a:has-text('FY25 SSA Contract Forecast')")
            
            if not forecast_links:
                logger.error("Could not find the FY25 SSA Contract Forecast link")
                return None
            
            # Click the link
            logger.info("Clicking the FY25 SSA Contract Forecast link")
            forecast_links[0].click()
            
            # Wait for the page to load
            logger.info("Waiting for page to load after clicking link")
            self.page.wait_for_selector("body", state="visible")
            
            # Look for the download link specifically for the 2025 file
            logger.info("Looking for the 2025 download link")
            # First try to find links with 01072025 in the text or href
            download_links = self.page.query_selector_all("a:has-text('01072025'), a[href*='01072025']")
            
            # If not found, look for links with 2025 in the text or href
            if not download_links:
                logger.info("No links with '01072025' found, looking for links with '2025'")
                download_links = self.page.query_selector_all("a:has-text('2025'), a[href*='2025']")
            
            # If still not found, fall back to any Excel/PDF links
            if not download_links:
                logger.info("No links with '2025' found, looking for Excel/PDF links")
                download_links = self.page.query_selector_all("a[href$='.xlsx'], a[href$='.xls'], a[href$='.pdf']")
            
            if not download_links:
                logger.error("Could not find any download links")
                return None
            
            # Click the download link and wait for the download to complete
            href = download_links[0].get_attribute('href')
            logger.info(f"Clicking download link: {href}")
            
            # Start waiting for the download
            with self.context.expect_download() as download_info:
                download_links[0].click()
            
            # Wait for the download to complete
            download = download_info.value
            logger.info(f"Download started: {download.suggested_filename}")
            
            # Wait for the download to complete and save to the downloads directory
            download_path = os.path.join(downloads_dir, download.suggested_filename)
            download.save_as(download_path)
            logger.info(f"Download completed: {download_path}")
            
            # Return the path to the downloaded file
            return download_path
        except Exception as e:
            logger.error(f"Error downloading forecast document: {e}")
            return None
    
    def process_excel(self, excel_file, session, data_source):
        """Process the Excel file and extract proposal data"""
        logger.info(f"Processing Excel file: {excel_file}")
        
        try:
            # First, try to convert the XLSM file to XLSX format
            import openpyxl
            import os
            
            # Create a new file path for the converted file
            file_dir = os.path.dirname(excel_file)
            file_name = os.path.basename(excel_file)
            file_base, file_ext = os.path.splitext(file_name)
            xlsx_file = os.path.join(file_dir, f"{file_base}_converted.xlsx")
            
            try:
                logger.info(f"Attempting to convert {excel_file} to {xlsx_file}")
                # Load the workbook
                wb = openpyxl.load_workbook(excel_file, read_only=False, data_only=True)
                # Save as xlsx
                wb.save(xlsx_file)
                logger.info(f"Successfully converted file to {xlsx_file}")
                # Use the converted file for processing
                excel_file = xlsx_file
            except Exception as e:
                logger.warning(f"Failed to convert file: {e}")
                # Continue with the original file
            
            # Try a more robust approach to read the Excel file
            try:
                # First try with engine='openpyxl' explicitly
                df = pd.read_excel(excel_file, engine='openpyxl')
            except Exception as e1:
                logger.warning(f"First attempt to read Excel file failed: {e1}")
                try:
                    # Try with xlrd engine for xls files
                    df = pd.read_excel(excel_file, engine='xlrd')
                except Exception as e2:
                    logger.warning(f"Second attempt to read Excel file failed: {e2}")
                    # Try with a different approach - read all sheets
                    try:
                        # Get all sheet names
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                        sheet_names = wb.sheetnames
                        logger.info(f"Available sheets: {sheet_names}")
                        
                        # Try to read the first sheet
                        df = pd.read_excel(excel_file, engine='openpyxl', sheet_name=sheet_names[0])
                    except Exception as e3:
                        # One last attempt - try to read the file as CSV
                        try:
                            logger.warning(f"Trying to read as CSV: {e3}")
                            df = pd.read_csv(excel_file)
                        except Exception as e4:
                            logger.error(f"All attempts to read Excel file failed: {e4}")
                            return 0
            
            # Check if the DataFrame is empty
            if df.empty:
                logger.error("Excel file is empty")
                return 0
            
            # Log the column names and first few rows for debugging
            logger.info(f"Column names: {df.columns.tolist()}")
            logger.info(f"First few rows:\n{df.head()}")
            
            # Try to find the header row by looking for specific keywords
            header_row_idx = None
            header_keywords = ['Site Type', 'App #', 'Requirement Type', 'Description', 'Est. Cost']
            
            # Look through the first 10 rows to find a potential header row
            for i in range(min(10, len(df))):
                row_values = df.iloc[i].astype(str).tolist()
                row_text = ' '.join(row_values).lower()
                
                # Check if this row contains multiple header keywords
                matches = sum(1 for keyword in header_keywords if keyword.lower() in row_text)
                if matches >= 3:  # If at least 3 keywords match, consider this the header row
                    header_row_idx = i
                    logger.info(f"Found header row at index {header_row_idx}")
                    break
            
            # If we couldn't find a header row, use a default (row 3 based on previous code)
            if header_row_idx is None:
                header_row_idx = 3
                logger.info(f"Using default header row at index {header_row_idx}")
            
            # Get the header row values
            header_values = df.iloc[header_row_idx].values
            
            # Create a mapping of column indices to header names
            col_mapping = {}
            for i, val in enumerate(header_values):
                if pd.notna(val) and str(val).strip():
                    col_mapping[i] = str(val).strip()
            
            # Create a new DataFrame with the data below the header row
            data_df = df.iloc[header_row_idx+1:].reset_index(drop=True)
            
            # Rename columns based on the mapping
            new_columns = [col_mapping.get(i, f"Unnamed_{i}") for i in range(len(data_df.columns))]
            data_df.columns = new_columns
            
            # Log the new column names
            logger.info(f"New column names: {data_df.columns.tolist()}")
            
            # Count of proposals added
            proposals_added = 0
            
            # Process each row
            for index, row in data_df.iterrows():
                try:
                    # Skip rows with no data
                    if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
                        continue
                    
                    # Extract values safely using iloc for positional access
                    # This avoids the ambiguity with duplicate column names
                    site_type = str(row.iloc[0]) if pd.notna(row.iloc[0]) else None
                    app_num = str(row.iloc[1]) if pd.notna(row.iloc[1]) else None
                    requirement_type = str(row.iloc[2]) if pd.notna(row.iloc[2]) else None
                    description = str(row.iloc[3]) if pd.notna(row.iloc[3]) else None
                    est_cost = str(row.iloc[4]) if pd.notna(row.iloc[4]) else None
                    award_date_str = str(row.iloc[5]) if pd.notna(row.iloc[5]) else None
                    existing_award = str(row.iloc[6]) if pd.notna(row.iloc[6]) else None
                    contract_type = str(row.iloc[7]) if pd.notna(row.iloc[7]) else None
                    incumbent = str(row.iloc[8]) if pd.notna(row.iloc[8]) else None
                    naics_code = str(row.iloc[9]) if pd.notna(row.iloc[9]) else None
                    naics_desc = str(row.iloc[10]) if pd.notna(row.iloc[10]) else None
                    competition_type = str(row.iloc[11]) if pd.notna(row.iloc[11]) else None
                    obligated_amt = str(row.iloc[12]) if pd.notna(row.iloc[12]) else None
                    place_of_performance = str(row.iloc[13]) if pd.notna(row.iloc[13]) else None
                    completion_date = str(row.iloc[14]) if pd.notna(row.iloc[14]) and index < len(data_df.columns) else None
                    
                    # Skip rows without a description
                    if not description or description.lower() in ['nan', 'none', '']:
                        continue
                    
                    # Parse dates
                    release_date = None
                    response_date = None
                    award_date = self.parse_date(award_date_str) if award_date_str and award_date_str.lower() not in ['nan', 'none', ''] else None
                    
                    # Parse estimated value
                    estimated_value = self.parse_value(est_cost) if est_cost and est_cost.lower() not in ['nan', 'none', ''] else None
                    
                    # Create a unique external ID
                    external_id = f"SSA_{app_num}" if app_num and app_num.lower() not in ['nan', 'none', ''] else None
                    
                    # Check if this proposal already exists
                    existing_proposal = None
                    if external_id:
                        existing_proposal = session.query(Proposal).filter(
                            Proposal.external_id == external_id,
                            Proposal.source_id == data_source.id,
                            Proposal.is_latest == True
                        ).first()
                    
                    if existing_proposal:
                        # Check if we need to update any fields that were previously empty or N/A
                        fields_updated = []
                        needs_update = False
                        
                        # Define field mappings for easier comparison
                        field_mappings = {
                            "title": description,
                            "office": site_type,
                            "description": description,
                            "naics_code": naics_code,
                            "estimated_value": estimated_value,
                            "release_date": release_date,
                            "response_date": response_date,
                            "url": self.base_url,
                            "contract_type": contract_type,
                            "competition_type": competition_type,
                            "solicitation_number": app_num,
                            "award_date": award_date,
                            "place_of_performance": place_of_performance,
                            "incumbent": incumbent
                        }
                        
                        # Check each field for updates
                        for field, new_value in field_mappings.items():
                            existing_value = getattr(existing_proposal, field)
                            
                            # Skip if both values are None
                            if existing_value is None and new_value is None:
                                continue
                            
                            # Check if we're updating from empty/N/A to a valid value
                            is_empty_or_na = (
                                existing_value is None or 
                                existing_value == "" or 
                                (isinstance(existing_value, str) and existing_value.lower() in ["n/a", "na", "not available", "not applicable", "tbd", "to be determined"])
                            )
                            
                            has_new_value = (
                                new_value is not None and 
                                new_value != "" and 
                                (not isinstance(new_value, str) or new_value.lower() not in ["n/a", "na", "not available", "not applicable", "tbd", "to be determined"])
                            )
                            
                            # Compare values, accounting for different types
                            if isinstance(existing_value, datetime.datetime) and isinstance(new_value, datetime.datetime):
                                # For dates, compare only the date part
                                if existing_value.date() != new_value.date():
                                    # If we're updating from a null-like date to a valid date
                                    if existing_value.year < 1900 and new_value.year >= 1900:
                                        setattr(existing_proposal, field, new_value)
                                        fields_updated.append(field)
                                        needs_update = True
                            elif existing_value != new_value:
                                # If we're updating from empty/N/A to a valid value
                                if is_empty_or_na and has_new_value:
                                    setattr(existing_proposal, field, new_value)
                                    fields_updated.append(field)
                                    needs_update = True
                        
                        if needs_update:
                            # Update the last_updated timestamp
                            existing_proposal.last_updated = datetime.datetime.utcnow()
                            session.add(existing_proposal)
                            logger.info(f"Updated existing proposal {external_id} - Fields updated: {fields_updated}")
                            proposals_added += 1
                        else:
                            logger.debug(f"No updates needed for existing proposal {external_id}")
                    else:
                        # Create a new proposal
                        proposal = Proposal(
                            source_id=data_source.id,
                            external_id=external_id,
                            title=description,
                            agency="Social Security Administration",
                            office=site_type,
                            description=description,
                            naics_code=naics_code,
                            estimated_value=estimated_value,
                            release_date=release_date,
                            response_date=response_date,
                            contact_info=None,
                            url=self.base_url,
                            status=None,
                            contract_type=contract_type,
                            set_aside=None,
                            competition_type=competition_type,
                            solicitation_number=app_num,
                            award_date=award_date,
                            place_of_performance=place_of_performance,
                            incumbent=incumbent,
                            is_latest=True
                        )
                        
                        session.add(proposal)
                        logger.info(f"Added new proposal: {description} (ID: {external_id})")
                        proposals_added += 1
                    
                except Exception as e:
                    logger.error(f"Error processing row {index}: {e}")
                    continue
            
            # Commit the changes
            session.commit()
            logger.info(f"Added {proposals_added} proposals from Excel file")
            
            return proposals_added
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            return 0
    
    def process_pdf(self, pdf_file, session, data_source):
        """Process the PDF file and extract proposal data"""
        logger.info(f"Processing PDF file: {pdf_file}")
        
        try:
            # For PDF processing, we would need to use a library like PyPDF2 or pdfplumber
            # This is a placeholder for now
            logger.warning("PDF processing is not implemented yet")
            return 0
            
        except Exception as e:
            logger.error(f"Error processing PDF file: {e}")
            return 0
    
    def cleanup_downloads(self):
        """Remove any non-2025 files from the downloads directory"""
        try:
            download_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'downloads')
            files = glob.glob(os.path.join(download_dir, "*"))
            
            # Find files that specifically contain "11042024" in their name (the non-2025 files)
            # and were modified in the last hour
            current_time = time.time()
            one_hour_ago = current_time - 3600  # 1 hour in seconds
            
            for file_path in files:
                file_name = os.path.basename(file_path)
                file_mod_time = os.path.getmtime(file_path)
                
                # Check if the file was modified in the last hour and contains "11042024" in the name
                if file_mod_time > one_hour_ago and "11042024" in file_name:
                    # Check if it's an Excel or PDF file (to avoid deleting other important files)
                    if file_path.endswith('.xlsx') or file_path.endswith('.xls') or file_path.endswith('.xlsm') or file_path.endswith('.pdf') or file_path.endswith('.csv'):
                        logger.info(f"Removing non-2025 file: {file_path}")
                        try:
                            os.remove(file_path)
                        except Exception as e:
                            logger.error(f"Error removing file {file_path}: {e}")
        
        except Exception as e:
            logger.error(f"Error during cleanup: {e}")

    def scrape(self):
        """Main scraping function"""
        logger.info("Starting scrape")
        
        try:
            # Set up the browser
            if not self.setup_browser():
                logger.error("Failed to set up browser")
                return False
            
            # Get a database session
            session = get_session()
            
            # Check if the data source exists
            data_source = session.query(DataSource).filter(DataSource.name == self.source_name).first()
            
            # If the data source doesn't exist, create it
            if not data_source:
                logger.info(f"Creating new data source: {self.source_name}")
                data_source = DataSource(
                    name=self.source_name,
                    url=self.base_url,
                    description="SSA Contract Forecast data source"
                )
                session.add(data_source)
                session.commit()
            
            # Download the forecast document
            forecast_file = self.download_forecast_document()
            
            if not forecast_file:
                logger.error("Failed to download forecast document")
                close_session(session)
                self.cleanup_browser()
                return False
            
            # Check if the downloaded file is the 2025 file
            if not ("01072025" in os.path.basename(forecast_file) or "2025" in os.path.basename(forecast_file)):
                logger.warning(f"Downloaded file does not appear to be the 2025 file: {forecast_file}")
                
                # Try to find a 2025 file in the downloads directory
                download_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'downloads')
                files = glob.glob(os.path.join(download_dir, "*"))
                
                # Filter for files with 01072025 or 2025 in the name
                year_2025_files = [f for f in files if "01072025" in os.path.basename(f) or "2025" in os.path.basename(f)]
                
                if year_2025_files:
                    # Sort by modification time (most recent first)
                    year_2025_files.sort(key=os.path.getmtime, reverse=True)
                    forecast_file = year_2025_files[0]
                    logger.info(f"Found 2025 file in downloads directory: {forecast_file}")
                else:
                    logger.warning("Could not find any 2025 files in the downloads directory")
            
            # Process the forecast document based on its type
            proposals_added = 0
            if forecast_file.endswith('.xlsx') or forecast_file.endswith('.xls') or forecast_file.endswith('.xlsm'):
                proposals_added = self.process_excel(forecast_file, session, data_source)
            elif forecast_file.endswith('.pdf'):
                proposals_added = self.process_pdf(forecast_file, session, data_source)
            else:
                logger.error(f"Unsupported file type: {forecast_file}")
            
            # Update the last scraped timestamp
            data_source.last_scraped = datetime.datetime.utcnow()
            session.commit()
            
            # Update the download tracker
            download_tracker.set_last_download_time(self.source_name)
            
            logger.info(f"Scrape completed. Added {proposals_added} proposals.")
            
            # Clean up any non-2025 files
            self.cleanup_downloads()
            
            # Close the session
            close_session(session)
            
            # Clean up the browser
            self.cleanup_browser()
            
            return True
            
        except Exception as e:
            logger.error(f"Error during scrape: {e}")
            
            # Clean up
            if 'session' in locals() and session:
                close_session(session)
            
            self.cleanup_browser()
            
            return False

def run_scraper(force=False):
    """Run the scraper if it's time or if forced"""
    try:
        # Get the scrape interval from environment or use default (24 hours)
        scrape_interval_hours = int(os.getenv("SCRAPE_INTERVAL_HOURS", 24))
        
        # Check if we should run the scraper using the download tracker
        if force or download_tracker.should_download("SSA Contract Forecast", scrape_interval_hours):
            logger.info(f"Running scraper (force={force})")
            
            # Create and run the scraper
            scraper = SSAContractForecastScraper(debug_mode=os.getenv("DEBUG", "False").lower() == "true")
            return scraper.scrape()
        else:
            logger.info(f"Skipping scrape (interval={scrape_interval_hours} hours)")
            return True
            
    except Exception as e:
        logger.error(f"Error running scraper: {e}")
        return False

if __name__ == "__main__":
    # Run the scraper with force=True to run regardless of when it was last run
    run_scraper(force=True) 